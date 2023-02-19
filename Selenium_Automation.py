import re
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
from csv import reader
global state, pin_code, street_Address

# driver for chrome
driver = webdriver.Chrome(executable_path='../drivers/chromedriver.exe')


def address_split(address, row1):
    # used to split address accordingly based on first website where it comes in 2 rows.
    global state, pin_code, street_Address
    if address == "" or address == "*":
        # if address is found to be empty tries to search in the second website.
        search_in_web2(str3)
    else:
        add = address.split('\n')
        street_Address = add[0]
        next_line = add[1]
        list1 = next_line.split()
        string1 = list1[1]
        for m in range(1, len(list1)):
            if len(list1[m]) > 2:
                string1.join(list1[m] + ' ')
            else:
                state = list1[m]
        pin_code = list1[len(list1) - 1]
        input_address(name_sheet2, row1)


def input_address(sheet, row1):
    # writes the address into the sheet accordingly.
    sheet['C'+str(row1)].value = street_Address
    sheet['D'+str(row1)].value = state
    sheet['E'+str(row1)].value = pin_code
    return 1


def search_in_web(string):
    # chooses search criteria.
    driver.get("https://businesssearch.sos.ca.gov/")
    # getting the business name as string and checking for LP or LLC word
    if string.find('LP') != -1 or string.find('LLC') != -1:
        driver.find_element_by_id("LLCNameOpt").click()
    else:
        driver.find_element_by_id("CorpNameOpt").click()
    # enters the string into the box.
    search_box = driver.find_element_by_class_name("form-control")
    search_box.send_keys(string)
    search_box.send_keys(Keys.RETURN)
    #  calculating the number of rows in the table.
    row_no = len(driver.find_elements_by_xpath("//*[@id='enitityTable']/tbody/tr"))
    if row_no == 0:
        # if no results , tries for second website.
        search_in_web2(str3)
        return 0
    before_XPath_1 = "//*[@id='enitityTable']/tbody/tr["
    after_XPath = "]/td[4]"
    # iterating through the rows of the table.
    for t_col in range(1, row_no):
        FinalXPath = before_XPath_1 + str(t_col) + after_XPath
        cell_text = driver.find_element_by_xpath(FinalXPath).text
        if cell_text.upper() == string.upper():
            # if string found in table further looks for address.
            driver.find_element_by_xpath(FinalXPath).click()
            address = driver.find_element_by_xpath('//*[@id="maincontent"]/div[3]/div[1]/div[7]/div[2]').text
            # gets the address and sends it to split function.
            address_split(address, row)
            return 0
    # sends for search to second website in case no results match.
    search_in_web2(str3)


def address_split2(address, row1):
    # used to split address accordingly.
    global state, pin_code, street_Address
    if address == "" or address == "*":
        # if address is empty returns empty strings.
        street_Address = ' '
        state = ' '
        pin_code = ' '
    else:
        # splits address.
        add = address.split(',')
        street_Address = add[0]
        next_line = add[1]
        list1 = next_line.split()
        state = list1[0]
        pincode = list1[1]
        input_address(name_sheet2, row1)


def search_in_web2(string):
    # search the second website.
    driver.get("https://www.corporationwiki.com/")
    search_box = driver.find_element_by_xpath("//*[@id='keywords']")
    search_box.send_keys(string)
    search_box.send_keys(Keys.RETURN)
    cell_text = driver.find_element_by_xpath("// *[ @ id = 'results-stats']").text
    list_no = cell_text.split()
    row_no = list_no[0]
    # if no results, end the search.
    if row_no == 0:
        return 0
    before_XPath_1 = "// *[ @ id = 'results-details'] / div["
    after_XPath = "] / div / div[1]"
    # iterating through the rows of the table.
    i = 0
    for t_col in range(2, 2 + int(row_no)):
        FinalXPath = before_XPath_1 + str(t_col) + after_XPath
        cell_text = driver.find_element_by_xpath(FinalXPath).text
        add1 = cell_text.split('\n')
        if add1[0].upper() == string.upper():
            # if matching then search for address.
            driver.find_element_by_xpath(FinalXPath+'/a').click()
            address = driver.find_element_by_xpath('//*[@id="main-content"]/div[7]/div/div/div/a/span').text
            address_split2(address, row)
            return 1
    if i == int(row_no):
        return 0


# opens workbook and deletes the columns.
workbook = load_workbook(filename="samplefile.xlsx")
name_sheet = workbook.active
name_sheet.delete_cols(2)
name_sheet.delete_cols(13)
name_sheet.delete_cols(13)
name_sheet.delete_cols(13)
workbook.save('samplefile2.xlsx')
# iterates through the table for business names.
row_count = name_sheet.max_row
workbook1 = load_workbook(filename="samplefile2.xlsx")
name_sheet2 = workbook1.active
list_row = []
for row in range(3, name_sheet2.max_row + 1):
    no = 'A' + str(row)
    str3 = name_sheet2[no].value
    str1 = str(str3).upper()
    with open('corp_terms_list.csv', 'r') as file:
        csv_reader = reader(file)
        i = 1
        length = len(pd.read_csv('corp_terms_list.csv'))

        for word in csv_reader:
            str2 = word[0].upper()
            i = i + 1
            if str2 in re.split(r"[\b\W\b]+ ", str1):
                break
            if str2 in str1.split():
                break

        if i == length + 2:
            list_row.append(row)
            workbook1.save('samplefile2.xlsx')
        else:
            search_in_web(str3)

for i in list_row:
    name_sheet2.delete_rows(i)

driver.quit()
workbook1.save('samplefile2.xlsx')







